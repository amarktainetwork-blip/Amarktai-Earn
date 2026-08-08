from __future__ import annotations

import csv
import json
import os
from pathlib import Path
from typing import Any, Iterable


class TabularError(ValueError):
    pass


def safe_spreadsheet_value(value: Any) -> Any:
    """Neutralize formula injection while preserving ordinary numbers and booleans."""
    if not isinstance(value, str):
        return value
    cleaned = value.replace("\x00", "")
    if cleaned.startswith(("=", "+", "-", "@", "\t", "\r")):
        return "'" + cleaned
    return cleaned


def _bounds(rows: list[dict[str, Any]]) -> tuple[int, int]:
    max_rows = max(1, int(os.getenv("TABULAR_MAX_ROWS", "100000")))
    max_columns = max(1, int(os.getenv("TABULAR_MAX_COLUMNS", "256")))
    columns = len({str(key) for row in rows for key in row})
    if len(rows) > max_rows:
        raise TabularError("TABULAR_ROW_LIMIT")
    if columns > max_columns:
        raise TabularError("TABULAR_COLUMN_LIMIT")
    return len(rows), columns


def normalize_rows(rows: Iterable[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    normalized: list[dict[str, Any]] = []
    headers: list[str] = []
    for raw in rows:
        if not isinstance(raw, dict):
            raise TabularError("tabular rows must be objects")
        row: dict[str, Any] = {}
        for raw_key, value in raw.items():
            key = str(safe_spreadsheet_value(str(raw_key).strip()))
            if not key:
                raise TabularError("tabular column names must not be blank")
            if key not in headers:
                headers.append(key)
            row[key] = value.strip() if isinstance(value, str) else value
        normalized.append(row)
    _bounds(normalized)
    return normalized, headers


def read_rows(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
    elif suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(payload, list):
            raise TabularError("source JSON must be a list of objects")
        rows = payload
    elif suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        try:
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(iterator, ())]
            if not headers or any(not header for header in headers):
                raise TabularError("workbook headers must not be blank")
            rows = [dict(zip(headers, values)) for values in iterator]
        finally:
            workbook.close()
    else:
        raise TabularError(f"unsupported tabular source: {suffix or 'none'}")
    return normalize_rows(rows)


def write_csv(path: Path, rows: list[dict[str, Any]], headers: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: safe_spreadsheet_value(row.get(key, "")) for key in headers})


def write_json(path: Path, rows: list[dict[str, Any]], headers: list[str]) -> None:
    material = [{key: row.get(key) for key in headers} for row in rows]
    path.write_text(json.dumps(material, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")


def write_xlsx(path: Path, rows: list[dict[str, Any]], headers: list[str], *, title: str = "Data") -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31] or "Data"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        sheet.append([safe_spreadsheet_value(row.get(key, "")) for key in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(48, max(10, *(len(str(cell.value or "")) + 2 for cell in column)))
    if headers and rows:
        table = Table(displayName="DataTable", ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        sheet.add_table(table)
    workbook.save(path)


def write_rows(path: Path, rows: list[dict[str, Any]], headers: list[str]) -> None:
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        write_csv(path, rows, headers)
    elif suffix == ".json":
        write_json(path, rows, headers)
    elif suffix == ".xlsx":
        write_xlsx(path, rows, headers)
    else:
        raise TabularError(f"unsupported tabular output: {suffix or 'none'}")
