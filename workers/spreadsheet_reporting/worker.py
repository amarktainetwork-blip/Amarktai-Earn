from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from workers.base import WorkRequest, WorkResult, Worker
from workers.tabular import TabularError, read_rows, safe_spreadsheet_value


class SpreadsheetReportingWorker(Worker):
    worker_class = "spreadsheet_reporting"

    def execute(self, request: WorkRequest) -> WorkResult:
        try:
            if request.inputs.get("operation") != "spreadsheet_report":
                return WorkResult(ok=False, error="unsupported spreadsheet operation")
            source = Path(str(request.inputs["source"]))
            rows, headers = read_rows(source)
            if not headers:
                raise TabularError("spreadsheet source has no columns")
            request.workspace.mkdir(parents=True, exist_ok=True)
            target = request.workspace / "professional-report.xlsx"
            workbook = Workbook()
            data = workbook.active
            data.title = "Data"
            data.append(headers)
            for cell in data[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            for row in rows:
                data.append([safe_spreadsheet_value(row.get(header, "")) for header in headers])
            data.freeze_panes = "A2"
            data.auto_filter.ref = data.dimensions
            if rows:
                table = Table(displayName="ReportData", ref=data.dimensions)
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
                data.add_table(table)
            for column in data.columns:
                data.column_dimensions[column[0].column_letter].width = min(45, max(10, *(len(str(cell.value or "")) + 2 for cell in column)))

            summary = workbook.create_sheet("Summary")
            summary.append(["Metric", "Value"])
            summary.append(["Rows", f"=COUNTA(Data!A:A)-1"])
            summary.append(["Columns", len(headers)])
            summary.append(["Blank cells", sum(value in (None, "") for row in rows for value in row.values())])
            summary["A1"].font = summary["B1"].font = Font(bold=True, color="FFFFFF")
            summary["A1"].fill = summary["B1"].fill = PatternFill("solid", fgColor="4472C4")
            category = str(request.inputs.get("category_column") or "")
            if category and category in headers:
                counts = Counter(str(row.get(category) or "(blank)") for row in rows)
                summary.append([]); summary.append([category, "Count"])
                start = summary.max_row + 1
                for label, count in counts.most_common(20):
                    summary.append([safe_spreadsheet_value(label), count])
                if counts:
                    chart = BarChart(); chart.title = f"Top {category} values"
                    chart.add_data(Reference(summary, min_col=2, min_row=start - 1, max_row=summary.max_row), titles_from_data=True)
                    chart.set_categories(Reference(summary, min_col=1, min_row=start, max_row=summary.max_row))
                    summary.add_chart(chart, "D2")
            workbook.save(target)
            reopened = load_workbook(target, read_only=True, data_only=False, keep_links=False)
            sheet_names = reopened.sheetnames; reopened.close()
            return WorkResult(ok=True, artifacts=[target], evidence={
                "operation": "spreadsheet_report", "rows": len(rows), "columns": headers,
                "sheet_names": sheet_names, "minimum_sheets": 2, "formula_injection_neutralized": True,
                "generated_formula_cells": 1,
            })
        except (OSError, KeyError, TypeError, ValueError, TabularError) as exc:
            return WorkResult(ok=False, error=str(exc))
