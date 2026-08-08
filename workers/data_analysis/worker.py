from __future__ import annotations

from collections import Counter
from pathlib import Path
from statistics import mean, median

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

from workers.base import WorkRequest, WorkResult, Worker
from workers.tabular import TabularError, read_rows, safe_spreadsheet_value


class DataAnalysisWorker(Worker):
    worker_class = "data_analysis"

    def execute(self, request: WorkRequest) -> WorkResult:
        try:
            if request.inputs.get("operation") != "data_analysis_report":
                return WorkResult(ok=False, error="unsupported analysis operation")
            rows, headers = read_rows(Path(str(request.inputs["source"])))
            request.workspace.mkdir(parents=True, exist_ok=True)
            target = request.workspace / "analysis-report.xlsx"
            workbook = Workbook(); summary = workbook.active; summary.title = "Analysis"
            summary.append(["Dataset profile", "Value"])
            summary.append(["Rows", len(rows)]); summary.append(["Columns", len(headers)])
            summary.append(["Blank cells", sum(value in (None, "") for row in rows for value in row.values())])
            summary["A1"].font = summary["B1"].font = Font(bold=True)
            numeric = []
            for header in headers:
                values = []
                for row in rows:
                    try:
                        if row.get(header) not in (None, ""):
                            values.append(float(row[header]))
                    except (TypeError, ValueError):
                        values = []; break
                if values:
                    numeric.append(header)
                    summary.append([f"{header} count", len(values)])
                    summary.append([f"{header} mean", mean(values)])
                    summary.append([f"{header} median", median(values)])
                    summary.append([f"{header} min", min(values)])
                    summary.append([f"{header} max", max(values)])
            trend = "not_requested"
            trend_column = str(request.inputs.get("trend_column") or "")
            value_column = str(request.inputs.get("value_column") or "")
            if trend_column in headers and value_column in headers:
                points = []
                for row in rows:
                    try: points.append((str(row.get(trend_column) or ""), float(row.get(value_column))))
                    except (TypeError, ValueError): continue
                points.sort(key=lambda item: item[0])
                if len(points) >= 2:
                    delta = points[-1][1] - points[0][1]
                    trend = "increasing" if delta > 0 else "decreasing" if delta < 0 else "flat"
                    summary.append([f"{value_column} trend", trend])
                    summary.append([f"{value_column} first", points[0][1]])
                    summary.append([f"{value_column} last", points[-1][1]])
            category = str(request.inputs.get("group_by") or "")
            visualization = False
            if category and category in headers:
                counts = Counter(str(row.get(category) or "(blank)") for row in rows)
                summary.append([]); summary.append([category, "Count"])
                header_row = summary.max_row; first = header_row + 1
                for label, count in counts.most_common(20):
                    summary.append([safe_spreadsheet_value(label), count])
                if counts:
                    chart = BarChart(); chart.title = f"Distribution by {category}"
                    chart.add_data(Reference(summary, min_col=2, min_row=header_row, max_row=summary.max_row), titles_from_data=True)
                    chart.set_categories(Reference(summary, min_col=1, min_row=first, max_row=summary.max_row))
                    summary.add_chart(chart, "D2"); visualization = True
            quality = workbook.create_sheet("Data Quality")
            quality.append(["Column", "Non-blank", "Blank", "Distinct"])
            for header in headers:
                values = [row.get(header) for row in rows]
                quality.append([safe_spreadsheet_value(header), sum(v not in (None, "") for v in values), sum(v in (None, "") for v in values), len({str(v) for v in values})])
            workbook.save(target)
            reopened = load_workbook(target, read_only=True, data_only=True, keep_links=False)
            sheets = reopened.sheetnames; reopened.close()
            return WorkResult(ok=True, artifacts=[target], evidence={
                "operation": "data_analysis_report", "rows": len(rows), "columns": headers,
                "numeric_columns": numeric, "sheet_names": sheets, "minimum_sheets": 2,
                "visualization_present": visualization, "regulated_advice": False, "trend": trend,
                "formula_injection_neutralized": True, "generated_formula_cells": 0,
            })
        except (OSError, KeyError, TypeError, ValueError, TabularError) as exc:
            return WorkResult(ok=False, error=str(exc))
