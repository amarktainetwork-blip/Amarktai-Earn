from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook

from control.models import Execution, Job, JobScore, Marketplace, QAResult
from control.services.execution import ExecutionError
from control.ops import agents_snapshot
from control.services.acquisition_preflight import run_acquisition_preflight
from planning.models import RepositorySnapshot, WorkPlan
from planning.services import execute_work_plan, plan_awarded_job, stage_local_job_asset


class _Response:
    def __init__(self, *, status=200, body=b"", content_type="text/plain", url="https://example.com/"):
        self.status_code = status; self._body = body; self.url = url; self.encoding = "utf-8"
        self.headers = {"Content-Type": content_type, "Content-Length": str(len(body))}
        self.text = body.decode("utf-8", errors="replace")
    def iter_content(self, chunk_size=65536):
        yield self._body
    def close(self): pass


class _Session:
    def __init__(self): self.trust_env = True; self.calls = 0
    def __enter__(self): return self
    def __exit__(self, *args): return False
    def get(self, url, **kwargs):
        self.calls += 1
        if url.endswith("/robots.txt"):
            return _Response(body=b"User-agent: *\nAllow: /\n", url=url)
        return _Response(body=b"<html lang='en'><head><title>Public evidence</title></head><body><h1>Evidence</h1><p>Bounded public content.</p></body></html>", content_type="text/html", url=url)


class ExpandedWorkersIntegrationTests(TestCase):
    def setUp(self):
        self.market = Marketplace.objects.create(slug="expanded-workers", display_name="Expanded Workers")
        self.temp = tempfile.TemporaryDirectory(); self.root = Path(self.temp.name)
        self.uploads = self.root / "uploads"; self.jobs = self.root / "jobs"; self.repos = self.root / "repos"
        self.uploads.mkdir(); self.jobs.mkdir(); self.repos.mkdir()
        self.env = patch.dict(os.environ, {
            "AMARKTAI_UPLOAD_ROOT": str(self.uploads), "AMARKTAI_JOB_ROOT": str(self.jobs), "AMARKTAI_REPO_ROOT": str(self.repos),
            "AMARKTAI_MIN_FREE_DISK_BYTES": "1", "AMARKTAI_MIN_FREE_DISK_PERCENT": "0",
            "AMARKTAI_MIN_MEMORY_HEADROOM_BYTES": "1", "AMARKTAI_MAX_LOAD_PER_CPU": "100", "PUBLIC_WEB_DATA_ENABLED": "1",
        }, clear=False); self.env.start()

    def tearDown(self):
        self.env.stop(); self.temp.cleanup()

    def _job(self, external_id: str, operation: str, **payload):
        return Job.objects.create(
            marketplace=self.market, external_id=external_id, title=f"Execute {operation}", task_class="Professional Services",
            reward="25", state=Job.State.AWARDED, normalized_payload={"operation": operation, **payload},
        )

    def _asset(self, job, name: str, text: str, role="source"):
        path = self.uploads / f"{job.external_id}-{name}"; path.write_text(text, encoding="utf-8")
        return stage_local_job_asset(job_id=job.id, path=str(path), semantic_role=role)

    def _run(self, job, expected_worker: str):
        plan = plan_awarded_job(job.id)
        self.assertEqual((plan.status, plan.worker_class), (WorkPlan.Status.READY, expected_worker), plan.reason_codes)
        with patch("control.services.execution.require_admission"):
            plan = execute_work_plan(plan.id)
        self.assertEqual(plan.status, WorkPlan.Status.QA_PASSED)
        execution = Execution.objects.filter(job=job).latest("attempt")
        self.assertEqual(execution.status, "QA_PASSED")
        self.assertTrue(QAResult.objects.filter(execution=execution, passed=True).exists())
        return execution

    def test_advanced_tabular_spreadsheet_and_analysis_execute_with_reopen_qa(self):
        advanced = self._job("advanced", "tabular_deduplicate", output_format="xlsx", keys=["name"])
        self._asset(advanced, "data.json", json.dumps([{"name": "Alice", "note": "=2+2"}, {"name": "Alice", "note": "duplicate"}, {"name": "Bob", "note": "ok"}]))
        execution = self._run(advanced, "advanced_structured_data")
        workbook = load_workbook(execution.artifacts.get().path, read_only=True, data_only=False, keep_links=False)
        try:
            self.assertEqual(workbook.active.max_row, 3)
            self.assertFalse([cell for row in workbook.active.iter_rows() for cell in row if cell.data_type == "f"])
        finally: workbook.close()

        report = self._job("spreadsheet", "spreadsheet_report", category_column="team")
        self._asset(report, "report.csv", "team,value\nA,10\nA,20\nB,30\n")
        self._run(report, "spreadsheet_reporting")

        analysis = self._job("analysis", "data_analysis_report", group_by="team", trend_column="period", value_column="value")
        self._asset(analysis, "analysis.csv", "period,team,value\n2026-01,A,10\n2026-02,A,20\n2026-03,B,30\n")
        execution = self._run(analysis, "data_analysis")
        self.assertEqual(execution.result["worker_evidence"]["trend"], "increasing")

    def test_dashboard_registry_exposes_every_expanded_worker_and_disabled_public_truth(self):
        rows = agents_snapshot()["rows"]
        classes = {row["worker_class"] for row in rows}
        self.assertTrue({
            "advanced_structured_data", "spreadsheet_reporting", "data_analysis", "technical_documentation",
            "content_copy", "seo_audit", "presentations", "document_production", "public_web_data",
            "web_output", "defensive_code_review", "customer_support",
        }.issubset(classes))
        with patch.dict(os.environ, {"PUBLIC_WEB_DATA_ENABLED": "0"}, clear=False):
            public = next(row for row in agents_snapshot()["rows"] if row["worker_class"] == "public_web_data")
        self.assertFalse(public["production_enabled"])
        self.assertIn("PUBLIC_WEB_DATA_DISABLED", public["enablement_reason_codes"])

    def test_acquisition_preflight_fails_closed_for_disabled_public_and_unauthorized_defensive_work(self):
        public = self._job("public-preflight", "public_web_extract", url="https://example.com/", purpose="", authorization_confirmed=False, terms_permit=False)
        JobScore.objects.create(job=public, p_acquire="1", p_accept="1", p_payment="1", expected_profit="10", expected_minutes=10)
        with patch.dict(os.environ, {"PUBLIC_WEB_DATA_ENABLED": "0", "ACQUISITION_ENABLED_OPERATIONS": "public_web_extract"}, clear=False):
            result = run_acquisition_preflight(public, persist=False)
        self.assertIn("PUBLIC_WEB_DATA_DISABLED", result.reason_codes)
        self.assertIn("PUBLIC_WEB_POLICY_PROOF_REQUIRED", result.reason_codes)

        defensive = self._job("defensive-preflight", "defensive_code_review", authorization_confirmed=False, scope="")
        JobScore.objects.create(job=defensive, p_acquire="1", p_accept="1", p_payment="1", expected_profit="10", expected_minutes=10)
        with patch.dict(os.environ, {"ACQUISITION_ENABLED_OPERATIONS": "defensive_code_review"}, clear=False):
            result = run_acquisition_preflight(defensive, persist=False)
        self.assertIn("DEFENSIVE_REVIEW_AUTHORIZATION_REQUIRED", result.reason_codes)
        self.assertIn("DEFENSIVE_REVIEW_SCOPE_REQUIRED", result.reason_codes)
        self.assertIn("REPOSITORY_NOT_DECLARED", result.reason_codes)

    def test_structured_file_and_web_artifact_workers_execute_with_independent_qa(self):
        seo = self._job("seo", "seo_content_audit", target_keywords=["page", "missing phrase"])
        self._asset(seo, "page.html", "<html lang='en'><head><title>Useful supplied page title</title><meta name='description' content='A sufficiently useful page description for deterministic content audit validation and review.'></head><body><h1>Page</h1><img src='x.png'></body></html>")
        execution = self._run(seo, "seo_audit")
        self.assertEqual(execution.result["worker_evidence"]["missing_target_keyword_count"], 1)

        deck = self._job("deck", "presentation_create", title="Quarterly update", slides=[{"title": "Overview", "body": "Summary"}, {"title": "Results", "body": "Measured results"}])
        self._run(deck, "presentations")

        docx = self._job("docx", "docx_create", title="Runbook", content="# Runbook\n\nThis runbook contains operational steps and recovery instructions.")
        self._run(docx, "document_production")
        pdf = self._job("pdf", "pdf_create", title="Report", content="This is a polished report with enough source-grounded content for a production artifact.")
        self._run(pdf, "document_production")

        html = self._job("html", "static_html_create", title="Accessible page", sections=[{"heading": "Overview", "body": "Useful accessible static content."}])
        execution = self._run(html, "web_output")
        self.assertEqual(execution.artifacts.count(), 2)

    def test_professional_text_workers_use_common_lifecycle_and_never_claim_external_send(self):
        fake = SimpleNamespace(model="dynamic-test-model")
        output = "# Professional Deliverable\n\nThis source-grounded deliverable contains clear structure, constraints, and enough substantive detail for independent structural quality validation."
        with patch("workers.professional_text.worker.generate_text", return_value=(output, fake)):
            content = self._job("content", "content_package", content_type="article", brief="Write an evidence-grounded article without invented claims.")
            self._run(content, "content_copy")
            support = self._job("support", "support_content_package", support_content_type="reply_draft", brief="Draft a response acknowledging the supplied issue without claiming an account action.")
            execution = self._run(support, "customer_support")
            self.assertTrue(execution.result["worker_evidence"]["draft_only"])

            documentation = self._job("documentation", "technical_documentation", documentation_type="readme", requirements="Document the supplied repository accurately.")
            repository = self.repos / "documentation"; repository.mkdir(); (repository / "app.py").write_text("def main():\n    return 'ok'\n", encoding="utf-8")
            RepositorySnapshot.objects.create(
                job=documentation, repository_url="https://github.com/example/repo", owner="example", repository="repo",
                commit_sha="a" * 40, path=str(repository), file_count=1, total_bytes=30,
                status=RepositorySnapshot.Status.VERIFIED, verified_at=timezone.now(),
            )
            self._run(documentation, "technical_documentation")

    def test_public_web_and_defensive_review_require_explicit_policy_scope(self):
        public = self._job(
            "public-web", "public_web_extract", url="https://example.com/evidence", purpose="Extract supplied public evidence",
            authorization_confirmed=True, terms_permit=True,
        )
        with patch("workers.public_web_data.worker.socket.getaddrinfo", return_value=[(None, None, None, None, ("93.184.216.34", 443))]), patch("workers.public_web_data.worker.requests.Session", _Session):
            self._run(public, "public_web_data")

        defensive = self._job("defensive", "defensive_code_review", authorization_confirmed=True, scope="Supplied repository only")
        repository = self.repos / "defensive"; repository.mkdir(); (repository / "app.py").write_text("def ok():\n    return True\n", encoding="utf-8")
        RepositorySnapshot.objects.create(
            job=defensive, repository_url="https://github.com/example/defensive", owner="example", repository="defensive",
            commit_sha="b" * 40, path=str(repository), file_count=1, total_bytes=30,
            status=RepositorySnapshot.Status.VERIFIED, verified_at=timezone.now(),
        )
        self._run(defensive, "defensive_code_review")

        blocked = self._job("defensive-blocked", "defensive_code_review", authorization_confirmed=False, scope="")
        plan = plan_awarded_job(blocked.id)
        self.assertEqual(plan.status, WorkPlan.Status.BLOCKED)
        self.assertIn("DEFENSIVE_REVIEW_AUTHORIZATION_REQUIRED", plan.reason_codes)
        self.assertIn("DEFENSIVE_REVIEW_SCOPE_REQUIRED", plan.reason_codes)

    def test_execution_rejects_cross_job_asset_paths(self):
        owner = self._job("asset-owner", "tabular_convert", output_format="csv")
        asset = self._asset(owner, "private.json", "[{\"private\": \"customer-a\"}]")
        attacker = self._job("other-job", "content_package", content_type="article", brief="Unrelated")
        plan = WorkPlan.objects.create(
            job=attacker, worker_class="advanced_structured_data", operation="tabular_convert",
            input_spec={"operation": "tabular_convert", "source": asset.path, "output_format": "csv"},
            status=WorkPlan.Status.READY,
        )
        with patch("control.services.execution.require_admission"), self.assertRaisesMessage(ExecutionError, "not registered to this job"):
            execute_work_plan(plan.id)
        self.assertFalse(Execution.objects.filter(job=attacker, status="QA_PASSED").exists())
