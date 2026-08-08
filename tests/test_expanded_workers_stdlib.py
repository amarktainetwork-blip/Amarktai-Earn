from __future__ import annotations

import tempfile
import unittest
import json
from pathlib import Path

from openpyxl import load_workbook

from workers.registry import all_specs, operation_spec, registered_operations
from workers.base import WorkRequest
from workers.advanced_structured_data.worker import AdvancedStructuredDataWorker
from workers.qa.runtime import run_qa
from workers.tabular import normalize_rows, safe_spreadsheet_value, write_xlsx


class ExpandedWorkerContractTests(unittest.TestCase):
    def test_expanded_registry_has_unique_factories_operations_and_qa(self):
        specs = all_specs()
        self.assertEqual(len(specs), 21)
        self.assertEqual(len({row.worker_class for row in specs}), 21)
        self.assertEqual(len(registered_operations()), 38)
        expected = {
            "advanced_structured_data", "spreadsheet_reporting", "data_analysis", "technical_documentation",
            "content_copy", "seo_audit", "presentations", "document_production", "public_web_data",
            "web_output", "defensive_code_review", "customer_support",
        }
        self.assertTrue(expected.issubset({row.worker_class for row in specs}))
        for operation in registered_operations():
            spec = operation_spec(operation)
            self.assertTrue(spec.factory)
            self.assertTrue(spec.qa_profile)

    def test_spreadsheet_formula_injection_is_neutralized(self):
        rows, headers = normalize_rows([{"name": "=2+2", "@command": "+SUM(A1:A2)"}])
        self.assertEqual(headers, ["name", "'@command"])
        self.assertEqual(safe_spreadsheet_value("-1+2"), "'-1+2")
        with tempfile.TemporaryDirectory() as tmp:
            target = Path(tmp) / "safe.xlsx"
            write_xlsx(target, rows, headers)
            workbook = load_workbook(target, read_only=True, data_only=False, keep_links=False)
            try:
                values = list(workbook.active.values)
                self.assertEqual(values[1][0], "'=2+2")
                self.assertEqual(values[1][1], "'+SUM(A1:A2)")
                self.assertFalse([cell for row in workbook.active.iter_rows() for cell in row if cell.data_type == "f"])
            finally:
                workbook.close()

    def test_every_advanced_tabular_operation_produces_a_qa_passable_artifact(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp); source = root / "source.json"; other = root / "other.json"
            source.write_text(json.dumps([{"id": 1, "name": " Alice ", "score": 2}, {"id": 2, "name": "Bob", "score": 10}, {"id": 2, "name": "Bob", "score": 10}]), encoding="utf-8")
            other.write_text(json.dumps([{"id": 1, "team": "A"}, {"id": 2, "team": "B"}]), encoding="utf-8")
            cases = [
                ("tabular_convert", {"source": str(source), "output_format": "csv"}),
                ("tabular_normalize", {"source": str(source), "output_format": "xlsx"}),
                ("tabular_deduplicate", {"source": str(source), "keys": ["id"], "output_format": "json"}),
                ("tabular_merge_join", {"sources": [str(source), str(other)], "join_key": "id", "output_format": "xlsx"}),
                ("tabular_filter_sort", {"source": str(source), "filter_equals": {"name": "Bob"}, "sort_by": "score", "descending": True, "output_format": "json"}),
                ("tabular_column_map", {"source": str(source), "column_mapping": {"name": "customer_name"}, "output_format": "csv"}),
                ("tabular_schema_validate", {"source": str(source), "schema": {"id": {"type": "integer", "required": True}}}),
            ]
            for index, (operation, inputs) in enumerate(cases):
                with self.subTest(operation=operation):
                    result = AdvancedStructuredDataWorker().execute(WorkRequest(job_id="job", workspace=root / f"work-{index}", inputs={"operation": operation, **inputs}))
                    self.assertTrue(result.ok, result.error)
                    qa = run_qa("tabular", result.artifacts[0], result.evidence)
                    self.assertTrue(qa.passed, (qa.checks, qa.evidence))


if __name__ == "__main__":
    unittest.main()
